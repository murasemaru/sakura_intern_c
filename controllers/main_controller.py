from flask import render_template, request, send_from_directory
import os
import csv
import re
import time
import sqlite3
from models.user_model import UserModel
from openpyxl import load_workbook, Workbook

UPLOAD_FOLDER = 'uploads'
SQL_FOLDER = 'sqls'
TMP_FOLDER = 'tmp'
META_DB = os.path.join(SQL_FOLDER, 'metadata.sqlite3')

def delete_metadata_and_db(filename):
    """同名ファイルのmetadataとDBファイルを削除"""
    conn = sqlite3.connect(META_DB)
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM metadata WHERE filename=?", (filename,))
        conn.commit()
    finally:
        conn.close()
    db_name = safe_filename(filename) + '.sqlite3'
    db_path = os.path.join(SQL_FOLDER, db_name)
    if os.path.exists(db_path):
        os.remove(db_path)


def init_metadata_db():
    """metadataテーブルを初期化"""
    conn = sqlite3.connect(META_DB)
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS metadata (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                tablename TEXT NOT NULL
            )
        """)
        conn.commit()
    finally:
        conn.close()

def get_metadata_files():
    """metadataテーブルに登録されているファイル名一覧を取得"""
    conn = sqlite3.connect(META_DB)
    try:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT filename FROM metadata")
        rows = cur.fetchall()
        return [r[0] for r in rows]
    finally:
        conn.close()

def get_metadata_list_grouped():
    """metadataテーブルの内容をデータベースごとにまとめて取得"""
    conn = sqlite3.connect(META_DB)
    try:
        cur = conn.cursor()
        cur.execute("SELECT filename, tablename FROM metadata")
        rows = cur.fetchall()  # [(filename, tablename), ...]
        grouped = {}
        for filename, tablename in rows:
            db_name = os.path.splitext(filename)[0] + '.sqlite3'  # ファイル名からDB名
            if db_name not in grouped:
                grouped[db_name] = []
            grouped[db_name].append(tablename)
        return grouped  # {'metadata.sqlite3': ['table1', 'table2'], ...}
    finally:
        conn.close()



def index_controller():
    UserModel.init_db()
    init_metadata_db()
    files = get_metadata_files()
    dev = request.args.get('dev') == '1'
    metadata_grouped = get_metadata_list_grouped()  # データベースごとにまとめる
    return render_template('index.html', files=files, dev=dev, metadata_grouped=metadata_grouped)



def safe_filename(filename):
    # 半角英数字とアンダースコアのみ許可
    name, _ = os.path.splitext(filename)
    return re.sub(r'[^A-Za-z0-9_]', '_', name)


def insert_metadata(filename, tablename):
    """metadataテーブルにレコードを追加"""
    conn = sqlite3.connect(META_DB)
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO metadata (filename, tablename) VALUES (?, ?)",
            (filename, tablename)
        )
        conn.commit()
    finally:
        conn.close()


def upload_controller():
    file = request.files['file']
    if not file:
        return render_template('message.html', message='ファイルを選択してください。')

    filename = file.filename
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    db_name = safe_filename(filename) + '.sqlite3'
    db_path = os.path.join(SQL_FOLDER, db_name)

    if filename.endswith('.csv'):
        table_name = safe_filename(filename)
        with open(filepath, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            columns = next(reader)
            conn = sqlite3.connect(db_path)
            try:
                cur = conn.cursor()
                col_defs = ', '.join([f'{c} TEXT' for c in columns])
                cur.execute(f'CREATE TABLE IF NOT EXISTS {table_name} ({col_defs})')
                for row in reader:
                    values = [v if v != '' else None for v in row]
                    col_names = ', '.join(columns)
                    placeholders = ', '.join(['?' for _ in columns])
                    cur.execute(f'INSERT INTO {table_name} ({col_names}) VALUES ({placeholders})', values)
                conn.commit()
            finally:
                conn.close()

        # metadataへ挿入
        insert_metadata(filename, table_name)

        return render_template('message.html', message=f'アップロード＆SQLite変換完了！<br>DBファイル名: {db_name}')

    elif filename.endswith('.xlsx'):
        wb = load_workbook(filepath, data_only=True)
        conn = sqlite3.connect(db_path)
        try:
            cur = conn.cursor()
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows or not rows[0]:
                    continue
                columns = [str(c) for c in rows[0]]
                table_name = f"{safe_filename(filename)}_{safe_filename(sheet)}"
                col_defs = ', '.join([f'{c} TEXT' for c in columns])
                cur.execute(f'CREATE TABLE IF NOT EXISTS {table_name} ({col_defs})')
                for row in rows[1:]:
                    values = [v if v is not None and v != '' else None for v in row]
                    col_names = ', '.join(columns)
                    placeholders = ', '.join(['?' for _ in columns])
                    cur.execute(f'INSERT INTO {table_name} ({col_names}) VALUES ({placeholders})', values)
            conn.commit()
        finally:
            conn.close()

        # metadataへ挿入（シートごと）
        for sheet in wb.sheetnames:
            table_name = f"{safe_filename(filename)}_{safe_filename(sheet)}"
            insert_metadata(filename, table_name)

        return render_template('message.html', message=f'アップロード＆SQLite変換完了！<br>DBファイル名: {db_name}')

    else:
        return render_template('message.html', message='CSVまたはXLSXファイルのみ対応しています。')


def download_controller(filename):
    # filename例: sample.csv, sample.xlsx, sample_sheet1.csv など
    # 管理DBから該当テーブル名を取得
    conn = sqlite3.connect(META_DB)
    try:
        cur = conn.cursor()
        # ファイル名に対応するテーブルをすべて取得
        cur.execute("SELECT tablename FROM metadata WHERE filename=?", (filename,))
        tables = [row[0] for row in cur.fetchall()]
    finally:
        conn.close()

    if not tables:
        return render_template('message.html', message='該当ファイルのテーブル情報がありません')

    # DBファイル名を決定
    db_name = safe_filename(filename) + '.sqlite3'
    db_path = os.path.join(SQL_FOLDER, db_name)
    if not os.path.exists(db_path):
        return render_template('message.html', message='該当DBファイルが存在しません')

    # 拡張子で出力形式を判定
    ext = os.path.splitext(filename)[1].lower()
    tmp_basename = safe_filename(filename) + '_' + str(int(time.time()))
    if ext == '.csv':
        # 1つ目のテーブルをcsv出力
        table = tables[0]
        tmpfile = os.path.join(TMP_FOLDER, tmp_basename + '.csv')
        conn = sqlite3.connect(db_path)
        try:
            cur = conn.cursor()
            cur.execute(f'SELECT * FROM {table}')
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description]
            with open(tmpfile, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                writer.writerows(rows)
        finally:
            conn.close()
        resp = send_from_directory(TMP_FOLDER, os.path.basename(tmpfile), as_attachment=True)
        try:
            os.remove(tmpfile)
        except Exception:
            pass
        return resp
    elif ext == '.xlsx':
        # 複数テーブルを1つのxlsxに出力
        tmpfile = os.path.join(TMP_FOLDER, tmp_basename + '.xlsx')
        wb = Workbook()
        wb.remove(wb.active)  # デフォルトシート削除
        conn = sqlite3.connect(db_path)
        try:
            cur = conn.cursor()
            for table in tables:
                ws = wb.create_sheet(title=table)
                cur.execute(f'SELECT * FROM {table}')
                rows = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
                ws.append(columns)
                for row in rows:
                    ws.append(row)
            wb.save(tmpfile)
        finally:
            conn.close()
        resp = send_from_directory(TMP_FOLDER, os.path.basename(tmpfile), as_attachment=True)
        try:
            os.remove(tmpfile)
        except Exception:
            pass
        return resp
    else:
        return render_template('message.html', message='対応していない拡張子です')

def sql_execute_controller():
    db_name = request.form.get('dbname')  # HTMLから入力されたDB名
    sql = request.form.get('sql')

    db_path = os.path.join(SQL_FOLDER, db_name)
    sql_result = ""

    if not os.path.exists(db_path):
        sql_result = f"データベース '{db_name}' が存在しません。"
    else:
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute(sql)
            # SELECT文ならfetchall
            if sql.strip().lower().startswith('select'):
                rows = cur.fetchall()
                # 結果を見やすく文字列化
                sql_result = '\n'.join([str(r) for r in rows])
            else:
                conn.commit()
                rows = cur.fetchall()
                # 結果を見やすく文字列化
                # sql_result = '\n'.join([str(r) for r in rows])
                sql_result = f"SQL実行成功: {cur.rowcount} 行が影響を受けました。"
        except Exception as e:
            sql_result = f"SQL実行エラー: {e}"
        finally:
            conn.close()

    # index_controllerと同じ情報を渡す
    files = os.listdir('uploads')
    metadata = index_controller().context['metadata'] if hasattr(index_controller(), 'context') else []
    dev = True

    metadata_grouped = get_metadata_list_grouped()

    return render_template('index.html', files=files, dev=dev, metadata_grouped=metadata_grouped, sql_result=sql_result)